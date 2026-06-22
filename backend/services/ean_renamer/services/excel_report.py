from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from services.ean_renamer.models import RenamePlanItem

REPORT_HEADERS = ["Date Change", "Old Name", "New Name", "File Type", "Folder "]
REPORT_WIDTHS = {
    "A": 20,
    "B": 25.5703125,
    "C": 20.7109375,
    "D": 22.140625,
    "E": 23,
}


def write_excel_report(output_root: Path, items: list[RenamePlanItem]) -> Path:
    output_root.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    path = unique_report_path(output_root / f"rename-report-{timestamp}.xlsx")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(REPORT_HEADERS)

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    today = datetime.now().date()
    for item in items:
        sheet.append(
            [
                today,
                item.oldName,
                item.newName,
                item.extension.lstrip("."),
                folder_name_for_item(item),
            ]
        )

    for column, width in REPORT_WIDTHS.items():
        sheet.column_dimensions[column].width = width
    for row in range(2, sheet.max_row + 1):
        sheet.cell(row=row, column=1).number_format = "m/d/yyyy"

    workbook.save(path)
    return path


def unique_report_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    for index in range(2, 1000):
        candidate = path.with_name(f"{stem}-{index}{path.suffix}")
        if not candidate.exists():
            return candidate
    raise RuntimeError("Could not create a unique report file name")


_CATEGORY_FOLDER_NAMES = {
    "packshot": "Packshot",
    "lifestyle_human": "Human",
    "lifestyle_normal": "Normal",
    "artwork": "Artwork",
}


def folder_name_for_item(item: RenamePlanItem) -> str:
    if item.category and item.category in _CATEGORY_FOLDER_NAMES:
        return _CATEGORY_FOLDER_NAMES[item.category]
    parts = Path(item.outputRelativePath or item.newName).parts
    if not parts:
        return ""
    if parts[0] == "Lifestyle" and len(parts) > 1:
        return parts[1]
    return parts[0]
