"""EAN Sorter — core scanning and sorting logic.
Ported from PRODUCT_DATA_CLEANER/ean_sorter.py.
"""
from __future__ import annotations

import re
import shutil
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

EAN_PATTERN = re.compile(r"(?<!\d)(\d{13}|\d{8})(?!\d)")

REPORT_NAME = "EAN_report.xlsx"
NOT_FOUND = "not found"


def ean_checksum_ok(code: str) -> bool:
    digits = [int(c) for c in code]
    check = digits.pop()
    total = sum(d * (3 if i % 2 == 0 else 1) for i, d in enumerate(reversed(digits)))
    return (10 - total % 10) % 10 == check


def extract_ean(name: str, is_file: bool = True) -> str | None:
    text = Path(name).stem if is_file else name
    candidates = EAN_PATTERN.findall(text)
    if not candidates:
        return None
    valid = [c for c in candidates if ean_checksum_ok(c)]
    pool = valid or candidates
    pool.sort(key=len, reverse=True)
    return pool[0]


def scan(root: Path) -> list[tuple[Path, str, str]]:
    found: list[tuple[Path, str, str]] = []

    def walk(directory: Path, depth: int) -> None:
        for item in sorted(directory.iterdir(), key=lambda p: p.name.lower()):
            if item.name.startswith(".") or item.name == REPORT_NAME:
                continue
            if depth == 0 and item.is_dir() and EAN_PATTERN.fullmatch(item.name):
                continue
            ean = extract_ean(item.name, is_file=not item.is_dir())
            if item.is_dir():
                if ean:
                    found.append((item, ean, "folder"))
                else:
                    found.append((item, NOT_FOUND, "folder"))
                    walk(item, depth + 1)
            else:
                found.append((item, ean or NOT_FOUND, "file"))

    walk(root, 0)
    return found


COLUMN_WIDTHS = {"A": 18.14, "B": 13.43, "C": 13.57, "D": 12.29, "E": 13.43, "F": 15.86}
HEADERS = ["NUMBERING", "EAN", "NAME", "TYPE", "OLD FOLDER", "NEW FOLDER"]
REPORT_FONT = Font(name="Aptos Narrow", size=11)


def write_report(path: Path, rows: list[tuple[str, str, str, str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = REPORT_FONT

    for row_idx, (name, ean, file_type, old_folder, new_folder) in enumerate(rows, 2):
        values = [row_idx - 1, ean, name, file_type, old_folder, new_folder]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = REPORT_FONT

    wb.save(path)


def sort_and_report(root: Path, delete_empty: bool = False) -> dict:
    found = scan(root)
    by_ean: dict[str, list[Path]] = {}
    for item, ean, _kind in found:
        if ean != NOT_FOUND:
            by_ean.setdefault(ean, []).append(item)

    report_rows: list[tuple[str, str, str, str, str]] = []
    moved = 0
    renamed: list[dict[str, str]] = []

    for ean, ean_items in by_ean.items():
        target = root / ean
        target.mkdir(exist_ok=True)
        for item in ean_items:
            if not item.exists():
                continue
            dest = target / item.name
            if dest.exists():
                stem = item.stem if item.is_file() else item.name
                suffix = item.suffix if item.is_file() else ""
                counter = 1
                while dest.exists():
                    dest = target / f"{stem}_{counter}{suffix}"
                    counter += 1
                renamed.append({"from": item.name, "to": dest.name})
            old_folder = str(item.parent)
            shutil.move(str(item), str(dest))
            moved += 1
            file_type = item.suffix.lstrip(".").upper() if item.is_file() else "FOLDER"
            report_rows.append((item.name, ean, file_type, old_folder, str(target)))

    for item, ean, _kind in found:
        if ean == NOT_FOUND:
            file_type = item.suffix.lstrip(".").upper() if item.is_file() else "FOLDER"
            report_rows.append((item.name, NOT_FOUND, file_type, str(item.parent), ""))

    report_path = root / REPORT_NAME
    write_report(report_path, report_rows)

    empty = _empty_folders(root)
    deleted = 0
    if delete_empty:
        for directory in empty:
            if directory.exists() and not any(directory.iterdir()):
                directory.rmdir()
                deleted += 1
        empty = _empty_folders(root)

    return {
        "moved": moved,
        "renamed": renamed,
        "report": str(report_path),
        "emptyFolders": [str(p) for p in empty],
        "deletedEmptyFolders": deleted,
        "products": len(by_ean),
    }


def _empty_folders(root: Path) -> list[Path]:
    return [
        d for d in sorted(root.rglob("*"), reverse=True)
        if d.is_dir() and not EAN_PATTERN.fullmatch(d.name) and not any(d.iterdir())
    ]
