"""EAN Sorter v2 — Excel report writer."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

REPORT_FONT = Font(name="Aptos Narrow", size=11)
REPORT_NAME = "EAN_sort_report.xlsx"

HEADERS = [
    "#",
    "Image Name",
    "Detected Type",
    "Detected Value",
    "Matched EAN",
    "Matched Product",
    "Confidence",
    "Source Folder",
    "Destination Folder",
    "Status",
]

COLUMN_WIDTHS = {
    "A": 6,
    "B": 28,
    "C": 15,
    "D": 20,
    "E": 16,
    "F": 28,
    "G": 12,
    "H": 30,
    "I": 30,
    "J": 14,
}


def write_sort_report(
    path: Path,
    match_results: list[dict],
    move_log: list[dict],
) -> None:
    dest_lookup: dict[str, str] = {}
    for entry in move_log:
        dest_lookup[entry["image_name"]] = entry["destination"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Sort Report"

    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(name="Aptos Narrow", size=11, bold=True)

    for row_idx, result in enumerate(match_results, 2):
        candidates = result.get("candidates", [])
        selected = result.get("selected_index")
        status = result.get("status", "unmatched")

        if candidates and selected is not None and 0 <= selected < len(candidates):
            best = candidates[selected]
        elif candidates:
            best = candidates[0]
        else:
            best = {}

        tier = best.get("tier", "")
        detected_value = best.get("match_source", "")
        matched_ean = best.get("ean", "")
        matched_product = best.get("product_name", "")
        confidence = best.get("confidence", 0)
        image_name = result.get("image_name", "")
        dest_folder = dest_lookup.get(image_name, "")

        values = [
            row_idx - 1,
            image_name,
            tier.upper() if tier else "NONE",
            detected_value,
            matched_ean,
            matched_product or "",
            f"{confidence:.0%}" if confidence else "",
            result.get("source_folder", ""),
            dest_folder,
            status.upper(),
        ]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = REPORT_FONT

    wb.save(path)


def write_loose_report(path: Path, moved_items: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Loose Images"

    headers = ["#", "Original Name", "Destination"]
    widths = {"A": 6, "B": 30, "C": 50}

    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(name="Aptos Narrow", size=11, bold=True)

    for row_idx, item in enumerate(moved_items, 2):
        values = [row_idx - 1, item.get("original", ""), item.get("destination", "")]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = REPORT_FONT

    wb.save(path)
