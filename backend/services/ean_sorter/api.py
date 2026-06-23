"""EAN Sorter — FastAPI router.
Wraps the core scanning/sorting logic in HTTP endpoints.
"""
from __future__ import annotations

import base64
import mimetypes
import os
import re
import shutil
import uuid
from collections import Counter
from pathlib import Path

import pandas as pd
from fastapi import APIRouter, BackgroundTasks, File, HTTPException, UploadFile

from services.data_maintenance.job_store import JobStore
from services.data_maintenance.models import JobStatus

from . import core

router = APIRouter()
job_store = JobStore()

IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp", ".tif", ".tiff"}


@router.get("/health")
def health() -> dict:
    return {"status": "ok", "service": "ean-sorter"}


@router.post("/scan")
def scan_folder(payload: dict) -> dict:
    folder = payload.get("folder", "")
    if not folder:
        raise HTTPException(status_code=400, detail="folder is required")
    root = Path(folder).expanduser()
    if not root.is_dir():
        raise HTTPException(status_code=400, detail=f"Not a folder: {root}")

    found = core.scan(root)
    return {
        "ok": True,
        "folder": str(root),
        **_summarize(found),
        "gallery": _gallery(root),
        "reportRows": _report_rows_from_scan(found),
    }


@router.post("/sort")
def sort_folder(payload: dict) -> dict:
    folder = payload.get("folder", "")
    delete_empty = payload.get("deleteEmpty", False)
    if not folder:
        raise HTTPException(status_code=400, detail="folder is required")
    root = Path(folder).expanduser()
    if not root.is_dir():
        raise HTTPException(status_code=400, detail=f"Not a folder: {root}")

    found = core.scan(root)
    summary = _summarize(found)
    result = core.sort_and_report(root, delete_empty=delete_empty)

    report_path = Path(result["report"])
    return {
        "ok": True,
        "folder": str(root),
        **result,
        **summary,
        "gallery": _gallery(root),
        "reportRows": _read_report(report_path),
    }


@router.post("/delete-empty")
def delete_empty(payload: dict) -> dict:
    folder = payload.get("folder", "")
    if not folder:
        raise HTTPException(status_code=400, detail="folder is required")
    root = Path(folder).expanduser()
    if not root.is_dir():
        raise HTTPException(status_code=400, detail=f"Not a folder: {root}")

    deleted = 0
    for d in core._empty_folders(root):
        if d.exists() and not any(d.iterdir()):
            d.rmdir()
            deleted += 1
    return {"ok": True, "deleted": deleted, "emptyFolders": [str(p) for p in core._empty_folders(root)]}


@router.post("/report")
def get_report(payload: dict) -> dict:
    folder = payload.get("folder", "")
    if not folder:
        raise HTTPException(status_code=400, detail="folder is required")
    report = Path(folder).expanduser() / core.REPORT_NAME
    if not report.is_file():
        raise HTTPException(status_code=404, detail="Report has not been created yet.")
    return {"ok": True, "report": str(report), "rows": _read_report(report)}


@router.post("/report/open")
def open_report_file(payload: dict) -> dict:
    folder = payload.get("folder", "")
    if not folder:
        raise HTTPException(status_code=400, detail="folder is required")
    report = Path(folder).expanduser() / core.REPORT_NAME
    if not report.is_file():
        raise HTTPException(status_code=404, detail="Report has not been created yet.")
    os.startfile(report)
    return {"ok": True, "report": str(report)}


@router.post("/report/export")
def export_report(payload: dict) -> dict:
    folder = payload.get("folder", "")
    destination = payload.get("destination", "")
    if not folder or not destination:
        raise HTTPException(status_code=400, detail="folder and destination are required")
    report = Path(folder).expanduser() / core.REPORT_NAME
    if not report.is_file():
        raise HTTPException(status_code=404, detail="Report has not been created yet.")
    dest = Path(destination)
    if dest.suffix.lower() != ".xlsx":
        dest = dest.with_suffix(".xlsx")
    shutil.copy2(report, dest)
    return {"ok": True, "report": str(dest)}


@router.post("/categorize/create-folders")
def create_category_folders(payload: dict) -> dict:
    folder = payload.get("folder", "")
    categories = payload.get("categories", [])
    if not folder:
        raise HTTPException(status_code=400, detail="folder is required")
    if not categories:
        raise HTTPException(status_code=400, detail="categories list is required")
    root = Path(folder).expanduser()
    if not root.is_dir():
        raise HTTPException(status_code=400, detail=f"Not a folder: {root}")

    created = []
    for cat in categories:
        cat_dir = root / cat
        cat_dir.mkdir(exist_ok=True)
        created.append(str(cat_dir))
    return {"ok": True, "created": created, "count": len(created)}


@router.post("/categorize/move")
def move_to_category(payload: dict) -> dict:
    folder = payload.get("folder", "")
    category = payload.get("category", "")
    paths = payload.get("paths", [])
    if not folder or not category:
        raise HTTPException(status_code=400, detail="folder and category are required")
    if not paths:
        raise HTTPException(status_code=400, detail="paths list is required")
    root = Path(folder).expanduser()
    target = root / category
    if not target.is_dir():
        raise HTTPException(status_code=400, detail=f"Category folder does not exist: {target}")

    moved = 0
    errors = []
    for p in paths:
        src = Path(p)
        if not src.exists():
            errors.append(f"Not found: {p}")
            continue
        dest = target / src.name
        if dest.exists():
            stem = src.stem if src.is_file() else src.name
            suffix = src.suffix if src.is_file() else ""
            counter = 1
            while dest.exists():
                dest = target / f"{stem}_{counter}{suffix}"
                counter += 1
        shutil.move(str(src), str(dest))
        moved += 1
    return {"ok": True, "moved": moved, "errors": errors}


@router.post("/categorize/uncategorized")
def get_uncategorized(payload: dict) -> dict:
    folder = payload.get("folder", "")
    if not folder:
        raise HTTPException(status_code=400, detail="folder is required")
    root = Path(folder).expanduser()
    if not root.is_dir():
        raise HTTPException(status_code=400, detail=f"Not a folder: {root}")

    found = core.scan(root)
    items = []
    for item, ean, kind in found:
        if ean == core.NOT_FOUND and item.exists():
            items.append({
                "name": item.name,
                "path": str(item),
                "kind": kind,
                "type": item.suffix.lstrip(".").upper() if item.is_file() else "FOLDER",
                "oldFolder": str(item.parent),
                "thumbnail": _thumbnail_for_item(item),
            })
    return {"ok": True, "items": items, "count": len(items)}


@router.post("/categorize/read-status-file")
async def read_status_file(file: UploadFile = File(...)) -> dict:
    import tempfile
    tmp = Path(tempfile.mktemp(suffix=".xlsx"))
    content = await file.read()
    tmp.write_bytes(content)
    try:
        return _read_status_workbook(tmp, file.filename or tmp.name)
    finally:
        try:
            tmp.unlink()
        except OSError:
            pass


@router.post("/categorize/read-status-file-path")
def read_status_file_path(payload: dict) -> dict:
    file_path = payload.get("path", "")
    if not file_path:
        raise HTTPException(status_code=400, detail="path is required")

    source = Path(file_path).expanduser()
    if not source.is_file():
        raise HTTPException(status_code=404, detail=f"Status file not found: {source}")
    if source.suffix.lower() not in {".xlsx", ".xls"}:
        raise HTTPException(status_code=400, detail="Status file must be an .xlsx or .xls workbook")

    return _read_status_workbook(source, source.name)


def _read_status_workbook(path: Path, filename: str) -> dict:
    try:
        xls = pd.ExcelFile(path)
        df = pd.read_excel(xls, xls.sheet_names[0], dtype={"BARCODE": str})
        xls.close()
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not read status workbook: {exc}") from exc

    required = {"OT ARTICLE NUMBER", "BARCODE", "STATUS", "PRODUCT NAME"}
    if not required.issubset(set(df.columns)):
        raise HTTPException(status_code=400, detail=f"Missing columns. Required: {sorted(required)}")

    brand = filename.split("_Missing_Data_Status")[0] if "_Missing_Data_Status" in filename else "BRAND"

    products = []
    no_barcode = []
    barcode_counts: Counter[str] = Counter()

    for _, row in df.iterrows():
        code = row.get("OT ARTICLE NUMBER")
        barcode_raw = row.get("BARCODE")
        status = row.get("STATUS")
        name = row.get("PRODUCT NAME")

        if pd.isna(barcode_raw) or str(barcode_raw).strip() in ("", "nan", "NaN"):
            barcode = None
        else:
            barcode = _normalize_barcode(barcode_raw)

        product = {
            "code": str(code) if pd.notna(code) else "",
            "barcode": barcode,
            "status": str(status) if pd.notna(status) else "",
            "name": str(name) if pd.notna(name) else "",
        }
        products.append(product)

        if barcode is None:
            no_barcode.append(product)
        else:
            barcode_counts[barcode] += 1

    duplicates = {bc: cnt for bc, cnt in barcode_counts.items() if cnt > 1}
    duplicate_products = [p for p in products if p["barcode"] in duplicates] if duplicates else []

    statuses: dict[str, int] = {}
    for p in products:
        s = p["status"] or "Unknown"
        statuses[s] = statuses.get(s, 0) + 1

    return {
        "ok": True,
        "brand": brand,
        "total": len(products),
        "products": products,
        "statuses": statuses,
        "no_barcode": no_barcode,
        "no_barcode_count": len(no_barcode),
        "duplicates": duplicates,
        "duplicate_products": duplicate_products,
    }


def _normalize_barcode(value: object) -> str:
    text = str(value).strip()
    if re.fullmatch(r"\d+\.0+", text):
        return text.split(".", 1)[0]
    if re.fullmatch(r"\d+", text):
        return text
    try:
        numeric = float(text)
    except ValueError:
        return text
    return str(int(numeric)) if numeric.is_integer() else text


@router.post("/categorize/create-status-folders")
def create_status_folders(payload: dict) -> dict:
    return create_status_folders_sync(payload)


@router.post("/categorize/create-status-folders-job")
def create_status_folders_job(payload: dict, background_tasks: BackgroundTasks) -> dict:
    destination = payload.get("destination", "")
    products = payload.get("products", [])
    statuses = payload.get("statuses", [])
    brand = payload.get("brand", "BRAND")
    if not destination:
        raise HTTPException(status_code=400, detail="destination is required")
    if not products:
        raise HTTPException(status_code=400, detail="products list is required")
    if not statuses:
        raise HTTPException(status_code=400, detail="statuses list is required")

    job_id = str(uuid.uuid4())
    job = job_store.create_job(
        job_id=job_id,
        job_type="ean_sorter_status_folders",
        original_filename=f"{brand} status folders",
        input_path=destination,
    )
    summary = {
        "progress_percent": 0,
        "current_file": "Queued",
        "created_count": 0,
        "skipped_count": 0,
        "total_products": len(products),
    }
    job_store.update_job(job_id, status=JobStatus.running, summary=summary)
    background_tasks.add_task(execute_status_folders_job, job_id, payload)
    return job_store.get_job(job.id).model_dump(mode="json")


def execute_status_folders_job(job_id: str, payload: dict) -> None:
    def update(progress: dict) -> None:
        current = job_store.get_job(job_id).summary or {}
        current.update(progress)
        job_store.update_job(job_id, status=JobStatus.running, summary=current)

    try:
        result = create_status_folders_sync(payload, progress_callback=update)
        summary = job_store.get_job(job_id).summary or {}
        summary.update({
            "progress_percent": 100,
            "current_file": "Completed",
            "created_count": result["count"],
            "skipped_count": result["skipped_count"],
            "created": result["created"],
            "skipped": result["skipped"],
        })
        job_store.update_job(job_id, status=JobStatus.completed, output_path=payload.get("destination", ""), error="", summary=summary)
    except Exception as exc:
        summary = job_store.get_job(job_id).summary or {}
        summary.update({"current_file": "Failed"})
        job_store.update_job(job_id, status=JobStatus.failed, error=str(exc), summary=summary)


def create_status_folders_sync(payload: dict, progress_callback=None) -> dict:
    destination = payload.get("destination", "")
    products = payload.get("products", [])
    statuses = payload.get("statuses", [])
    brand = payload.get("brand", "BRAND")
    use_name_for_no_barcode = payload.get("use_name_for_no_barcode", False)
    no_barcode_statuses = {str(status).strip() for status in payload.get("no_barcode_statuses", [])}
    per_product_for_duplicates = payload.get("per_product_for_duplicates", False)

    if not destination:
        raise HTTPException(status_code=400, detail="destination is required")
    if not statuses:
        raise HTTPException(status_code=400, detail="statuses list is required")

    root = Path(destination).expanduser()
    if not root.is_dir():
        raise HTTPException(status_code=400, detail=f"Not a folder: {root}")

    created_folders = []
    skipped = []
    selected_statuses = set(statuses)
    selected_products = [p for p in products if p.get("status", "").strip() in selected_statuses]
    total = max(1, len(selected_products))
    completed = 0

    barcode_counter: Counter[str] = Counter()
    for p in products:
        bc = p.get("barcode")
        if bc:
            barcode_counter[bc] += 1
    duplicate_barcodes = {bc for bc, cnt in barcode_counter.items() if cnt > 1}

    for p in products:
        status = p.get("status", "").strip()
        if status not in selected_statuses:
            continue
        completed += 1

        barcode = p.get("barcode")
        name = p.get("name", "")
        code = p.get("code", "")

        status_dir = root / status
        status_dir.mkdir(exist_ok=True)

        if barcode and barcode not in duplicate_barcodes:
            sub = status_dir / barcode
            sub.mkdir(exist_ok=True)
            created_folders.append(str(sub))
        elif barcode and barcode in duplicate_barcodes:
            if per_product_for_duplicates:
                safe_name = re.sub(r'[<>:"/\\|?*]', '_', name.strip())
                folder_name = f"{brand}_{safe_name}_{status}"
                sub = status_dir / folder_name
                sub.mkdir(exist_ok=True)
                created_folders.append(str(sub))
            else:
                sub = status_dir / barcode
                sub.mkdir(exist_ok=True)
                if str(sub) not in created_folders:
                    created_folders.append(str(sub))
        elif not barcode:
            if use_name_for_no_barcode and (not no_barcode_statuses or status in no_barcode_statuses):
                safe_name = re.sub(r'[<>:"/\\|?*]', '_', name.strip())
                folder_name = f"{brand}_{safe_name}_{status}"
                sub = status_dir / folder_name
                sub.mkdir(exist_ok=True)
                created_folders.append(str(sub))
            else:
                skipped.append({"code": code, "name": name, "status": status})
        if progress_callback and (completed == total or completed % 50 == 0):
            progress_callback({
                "progress_percent": int(round(completed / total * 100)),
                "current_file": f"Creating {status} folders",
                "created_count": len(created_folders),
                "skipped_count": len(skipped),
                "total_products": len(selected_products),
            })

    return {
        "ok": True,
        "created": created_folders,
        "count": len(created_folders),
        "skipped": skipped,
        "skipped_count": len(skipped),
    }


@router.post("/reveal")
def reveal_folder(payload: dict) -> dict:
    folder = payload.get("folder", "")
    if not folder:
        raise HTTPException(status_code=400, detail="folder is required")
    root = Path(folder).expanduser()
    if not root.exists():
        raise HTTPException(status_code=404, detail=f"Path does not exist: {root}")
    os.startfile(root)
    return {"ok": True}


def _summarize(found: list[tuple[Path, str, str]]) -> dict:
    products: dict[str, int] = {}
    rows = []
    for item, ean, kind in found:
        if ean != core.NOT_FOUND:
            products[ean] = products.get(ean, 0) + 1
        rows.append({
            "name": item.name,
            "path": str(item),
            "ean": ean,
            "kind": kind,
            "type": item.suffix.lstrip(".").upper() if item.is_file() else "FOLDER",
            "oldFolder": str(item.parent),
            "thumbnail": _thumbnail_for_item(item),
        })
    return {
        "items": len(found),
        "files": sum(1 for _, _, k in found if k == "file"),
        "folders": sum(1 for _, _, k in found if k == "folder"),
        "notFound": sum(1 for _, e, _ in found if e == core.NOT_FOUND),
        "products": len(products),
        "rows": rows,
        "productRows": [{"ean": ean, "count": count} for ean, count in sorted(products.items())],
    }


def _report_rows_from_scan(found: list[tuple[Path, str, str]]) -> list[dict]:
    return [
        {
            "numbering": i,
            "ean": ean,
            "name": item.name,
            "type": item.suffix.lstrip(".").upper() if item.is_file() else "FOLDER",
            "oldFolder": str(item.parent),
            "newFolder": "",
        }
        for i, (item, ean, _) in enumerate(found, 1)
    ]


def _read_report(report: Path) -> list[dict]:
    if not report.is_file():
        return []
    from openpyxl import load_workbook
    wb = load_workbook(report, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(v is not None for v in row):
            continue
        rows.append({
            "numbering": row[0] or "",
            "ean": row[1] or "",
            "name": row[2] or "",
            "type": row[3] or "",
            "oldFolder": row[4] or "",
            "newFolder": row[5] or "",
        })
    wb.close()
    return rows


def _gallery(root: Path) -> list[dict]:
    images = []
    for img in sorted(root.rglob("*"), key=lambda p: str(p).lower()):
        if img.is_file() and img.suffix.lower() in IMAGE_SUFFIXES:
            images.append({
                "name": img.name,
                "path": str(img),
                "folder": str(img.parent),
                "ean": core.extract_ean(img.name, True) or core.NOT_FOUND,
                "thumbnail": _data_url(img),
            })
    return images


def _thumbnail_for_item(item: Path) -> str:
    if item.is_file() and item.suffix.lower() in IMAGE_SUFFIXES:
        return _data_url(item)
    if item.is_dir():
        for img in sorted(item.rglob("*"), key=lambda p: str(p).lower()):
            if img.is_file() and img.suffix.lower() in IMAGE_SUFFIXES:
                return _data_url(img)
    return ""


def _data_url(path: Path) -> str:
    try:
        mime = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
        encoded = base64.b64encode(path.read_bytes()).decode("ascii")
        return f"data:{mime};base64,{encoded}"
    except Exception:
        return ""
