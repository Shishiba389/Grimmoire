from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .auditor import AuditResult
from .field_mapping import OVERVIEW_FIELD_GROUPS


HEADER_BLUE = "1F4E79"
GROUP_COLORS = {
    "Basic product info": "C00000",
    "Consumer pack": "C00000",
    "Trade pack": "2E75B6",
    "Outer box": "2E75B6",
    "Registration": "7030A0",
    "Manufacturer & RP": "375623",
    "Compliance": "2E75B6",
    "Product price": "C00000",
    "Barcode": "C00000",
}
BASE_VISIBLE_SHEETS = {"Missing Data Overview", "Summary Tracker", "Action Tracker", "Run Summary"}
STATUS_COLUMNS = ["Brand", "Total", "Active", "Upcoming", "Limited", "Non-Active", "Discontinued", "Blanks", "N/A", "Unknown", "Non-ACR", "Others"]


def write_audit_report(result: AuditResult, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        write_missing_overview_sheet(writer, result)
        write_summary_tracker_sheet(writer, result)
        write_action_tracker_sheet(writer, result)
        write_update_summary_sheet(writer, result)
        write_run_summary_sheet(writer, result)
        write_generic_sheet(writer, "Brand Scorecard", result.brand_scorecard)
        write_generic_sheet(writer, "SKU Missing Detail", result.sku_missing_detail)
        write_generic_sheet(writer, "Validation Errors", result.validation_errors)

    format_workbook(output_path)
    return output_path


def non_empty_dataframe(dataframe: pd.DataFrame) -> pd.DataFrame:
    if dataframe.empty:
        return pd.DataFrame({"Message": ["No rows"]})
    return dataframe


def run_summary_value(result: AuditResult, item: str) -> str:
    if result.run_summary.empty:
        return ""
    rows = result.run_summary[result.run_summary["Item"] == item]
    if rows.empty:
        return ""
    return str(rows.iloc[0]["Value"])


def is_update_summary_sheet(sheet_name: str) -> bool:
    return sheet_name.startswith("Update Summary v")


def is_visible_sheet(sheet_name: str) -> bool:
    return sheet_name in BASE_VISIBLE_SHEETS or is_update_summary_sheet(sheet_name)


def write_missing_overview_sheet(writer: pd.ExcelWriter, result: AuditResult) -> None:
    sheet_name = "Missing Data Overview"
    dataframe = non_empty_dataframe(result.missing_overview)
    dataframe.to_excel(writer, index=False, sheet_name=sheet_name, startrow=4)
    worksheet = writer.sheets[sheet_name]
    worksheet.cell(1, 1).value = f"BRAND DATA COMPLETENESS REPORT - {result.summary.brand_count} BRANDS"
    worksheet.cell(2, 1).value = (
        f"Source: {run_summary_value(result, 'Source file')} | "
        "Status profile: default_v8 | "
        f"Included statuses: {run_summary_value(result, 'Included statuses') or 'configured rule profile'}"
    )
    fields_to_groups = dict(OVERVIEW_FIELD_GROUPS)
    for col_index, column in enumerate(dataframe.columns, start=1):
        worksheet.cell(4, col_index).value = fields_to_groups.get(str(column), None)
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, dataframe.shape[1]))


def write_summary_tracker_sheet(writer: pd.ExcelWriter, result: AuditResult) -> None:
    sheet_name = "Summary Tracker"
    dataframe = non_empty_dataframe(result.summary_tracker)
    dataframe.to_excel(writer, index=False, sheet_name=sheet_name, startrow=2)
    writer.sheets[sheet_name].cell(1, 1).value = "SUMMARY TRACKER"


def write_action_tracker_sheet(writer: pd.ExcelWriter, result: AuditResult) -> None:
    sheet_name = "Action Tracker"
    dataframe = non_empty_dataframe(result.action_tracker)
    expected_cols = ["#", "Brand", "Field", "# Missing", "Total", "% Missing", "Priority", "Status"]
    dataframe = dataframe[[column for column in expected_cols if column in dataframe.columns]]
    dataframe.to_excel(writer, index=False, sheet_name=sheet_name, startrow=2)
    writer.sheets[sheet_name].cell(1, 1).value = "ACTION TRACKER"


def write_generic_sheet(writer: pd.ExcelWriter, sheet_name: str, dataframe: pd.DataFrame) -> None:
    non_empty_dataframe(dataframe).to_excel(writer, index=False, sheet_name=sheet_name)


def write_update_summary_sheet(writer: pd.ExcelWriter, result: AuditResult) -> None:
    sheet_name = result.update_summary_sheet_name or "Update Summary v1"
    if not result.update_summary.empty:
        result.update_summary.to_excel(writer, index=False, header=False, sheet_name=sheet_name)
        worksheet = writer.sheets[sheet_name]
        worksheet.merge_cells("A1:D1")
        return
    source_file = run_summary_value(result, "Source file")
    rows = [
        ["UPDATE SUMMARY - Current Master Data Audit", "", "", ""],
        ["", "", "", ""],
        ["1. FILE CHANGES OVERVIEW", "", "", ""],
        ["Item", "Detail", "", ""],
        ["Master data", source_file, "", ""],
        ["Total rows", result.summary.total_rows, "", ""],
        ["Included rows", result.summary.included_rows, "", ""],
        ["Brand count", result.summary.brand_count, "", ""],
        ["", "", "", ""],
        ["2. FIELD MAPPING REFERENCE", "", "", ""],
        ["Report Field", "Source Column / Logic", "Group", "Notes"],
    ]
    rows.extend([[field, "Mapped from master data aliases", group, "Missing source column is shown as -"] for field, group in OVERVIEW_FIELD_GROUPS])
    rows.extend(
        [
            ["", "", "", ""],
            ["3. STATUS FILTER", "", "", ""],
            ["Included statuses", run_summary_value(result, "Included statuses") or "configured rule profile", "", ""],
            ["Rule", "Rows included in Total and missing counts must match selected statuses.", "", ""],
            ["", "", "", ""],
            ["4. VALIDATION SUMMARY", "", "", ""],
            ["Action rows", result.summary.action_count, "", ""],
            ["Critical actions", result.summary.critical_actions, "", ""],
            ["Validation errors", result.summary.validation_error_count, "", ""],
            ["Warnings", " | ".join(result.summary.warnings), "", ""],
        ]
    )
    pd.DataFrame(rows).to_excel(writer, index=False, header=False, sheet_name=sheet_name)
    worksheet = writer.sheets[sheet_name]
    worksheet.merge_cells("A1:D1")


def write_run_summary_sheet(writer: pd.ExcelWriter, result: AuditResult) -> None:
    rows = [
        ("Run Summary", ""),
        ("Master data", run_summary_value(result, "Source file")),
        ("Brand count", result.summary.brand_count),
        ("Included statuses", run_summary_value(result, "Included statuses") or "configured rule profile"),
        ("Errors", 0),
        ("Warnings", len(result.summary.warnings)),
        ("", ""),
        ("Validation messages", " | ".join(result.summary.warnings)),
    ]
    pd.DataFrame(rows).to_excel(writer, index=False, header=False, sheet_name="Run Summary")


def format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor=HEADER_BLUE)
    header_font = Font(color="FFFFFF", bold=True)
    white_bold = Font(color="FFFFFF", bold=True)
    ok_fill = PatternFill("solid", fgColor="C6EFCE")
    ok_font = Font(color="006100")
    missing_fill = PatternFill("solid", fgColor="FFC7CE")
    missing_font = Font(color="9C0006", bold=True)
    dash_fill = PatternFill("solid", fgColor="D9E1F2")
    dash_font = Font(color="666666", italic=True)
    thin_border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )
    for worksheet in workbook.worksheets:
        worksheet.sheet_view.showGridLines = False
        header_row = header_row_for_sheet(worksheet.title)
        worksheet.freeze_panes = freeze_pane_for_sheet(worksheet.title)
        if not is_visible_sheet(worksheet.title):
            worksheet.sheet_state = "hidden"
        for cell in worksheet[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        if worksheet.title == "Missing Data Overview":
            for cell in worksheet[4]:
                if cell.value:
                    cell.fill = PatternFill("solid", fgColor=GROUP_COLORS.get(str(cell.value), "808080"))
                    cell.font = white_bold
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
            worksheet.row_dimensions[4].height = 22
            worksheet.row_dimensions[5].height = 32
        if is_update_summary_sheet(worksheet.title):
            worksheet["A1"].fill = PatternFill("solid", fgColor="E2F0D9")
            worksheet["A1"].font = Font(color=HEADER_BLUE, bold=True, size=12)
            for row in worksheet.iter_rows():
                first = row[0].value
                if isinstance(first, str) and first[:2].isdigit() is False and first.endswith("OVERVIEW"):
                    row[0].font = Font(color=HEADER_BLUE, bold=True)
            for row_index in range(1, worksheet.max_row + 1):
                first_value = worksheet.cell(row_index, 1).value
                if isinstance(first_value, str) and first_value[:2] in {"1.", "2.", "3.", "4."}:
                    worksheet.cell(row_index, 1).font = Font(color=HEADER_BLUE, bold=True)
                if worksheet.cell(row_index, 1).value in {"Item", "Report Field"}:
                    for cell in worksheet[row_index]:
                        cell.fill = header_fill
                        cell.font = header_font
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="center")
                if cell.row >= header_row:
                    cell.border = thin_border
                if isinstance(cell.value, str):
                    if cell.value == "NO missing":
                        cell.fill = ok_fill
                        cell.font = ok_font
                    elif cell.value.startswith("missing:"):
                        cell.fill = missing_fill
                        cell.font = missing_font
                    elif cell.value == "-":
                        cell.fill = dash_fill
                        cell.font = dash_font
        apply_sample_column_widths(worksheet)
    workbook.save(path)


def header_row_for_sheet(sheet_name: str) -> int:
    if sheet_name == "Missing Data Overview":
        return 5
    if sheet_name in {"Summary Tracker", "Action Tracker"}:
        return 3
    return 1


def freeze_pane_for_sheet(sheet_name: str) -> str | None:
    if sheet_name == "Missing Data Overview":
        return "M6"
    if sheet_name == "Summary Tracker":
        return "M4"
    return None


def apply_sample_column_widths(worksheet) -> None:
    if worksheet.title == "Missing Data Overview":
        widths = {
            "A": 47, "B": 10, "C": 10, "D": 10, "E": 10, "F": 12, "G": 14, "H": 10, "I": 10, "J": 10, "K": 10, "L": 10,
            "M": 20, "N": 20, "O": 20, "P": 16, "Q": 18, "R": 16, "S": 15, "T": 15, "U": 15, "V": 20, "W": 20,
            "X": 14, "Y": 14, "Z": 15, "AA": 20, "AB": 20, "AC": 14, "AD": 14, "AE": 14, "AF": 16, "AG": 19,
            "AH": 22, "AI": 19, "AJ": 23, "AK": 23, "AL": 17, "AM": 14, "AN": 21, "AO": 26, "AP": 25, "AQ": 14,
            "AR": 14, "AS": 22, "AT": 27, "AU": 25, "AV": 15, "AW": 12, "AX": 15, "AY": 17, "AZ": 17, "BA": 13,
        }
    elif worksheet.title == "Summary Tracker":
        widths = {"A": 18, "B": 10, "C": 10, "D": 10, "E": 10, "F": 12, "G": 14, "H": 10, "I": 10, "J": 10, "K": 10, "L": 10}
        for column in range(13, worksheet.max_column + 1):
            widths[get_column_letter(column)] = 20
    elif worksheet.title == "Action Tracker":
        widths = {"A": 16, "B": 18, "C": 26, "D": 11, "E": 10, "F": 22, "G": 10, "H": 10}
    elif is_update_summary_sheet(worksheet.title):
        widths = {"A": 47, "B": 47, "C": 47, "D": 12}
    elif worksheet.title == "Run Summary":
        widths = {"A": 21, "B": 47}
    else:
        widths = {}

    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        if column_letter in widths:
            worksheet.column_dimensions[column_letter].width = widths[column_letter]
            continue
        max_length = 0
        for cell in column_cells[:200]:
            max_length = max(max_length, len(str(cell.value or "")))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)
