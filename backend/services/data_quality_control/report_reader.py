from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SHEET_HEADER_ROWS = {
    "Missing Data Overview": 5,
    "Summary Tracker": 3,
    "Action Tracker": 3,
    "Brand Scorecard": 1,
    "SKU Missing Detail": 1,
    "Validation Errors": 1,
    "Run Summary": 1,
}


def read_report_workbook(path: Path, max_rows_per_sheet: int = 5000) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets: dict[str, Any] = {}
    for worksheet in workbook.worksheets:
        header_row = SHEET_HEADER_ROWS.get(worksheet.title, 1)
        headers = [
            str(value).strip() if value is not None and str(value).strip() else f"Column {index}"
            for index, value in enumerate(
                next(worksheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True)),
                start=1,
            )
        ]
        rows: list[dict[str, Any]] = []
        for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
            if len(rows) >= max_rows_per_sheet:
                break
            if not any(value is not None and str(value).strip() for value in row):
                continue
            rows.append({headers[index]: normalize_cell(value) for index, value in enumerate(row[: len(headers)])})
        sheets[worksheet.title] = {
            "headers": headers,
            "rows": rows,
            "row_count": len(rows),
            "returned_count": len(rows),
        }
    return {"path": str(path), "sheets": sheets}


def normalize_cell(value: Any) -> Any:
    if value is None:
        return ""
    return value
