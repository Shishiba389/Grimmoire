from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from .field_mapping import FIELD_ALIASES


CANONICAL_ALIASES = {
    "brand": ["brand", "brand name", "brand_name", "BRAND", "OBJECT"],
    "status": ["status", "product status", "item status", "STATUS"],
    "sku": ["sku", "ot article number", "article number", "item code", "product code", "CODE"],
    "product_name": ["product name", "name", "item name", "description", "NAME"],
}


@dataclass(frozen=True)
class ReadResult:
    dataframe: pd.DataFrame
    sheet_name: str
    header_row: int
    source_path: Path
    columns: list[str]
    warnings: list[str]


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\n", " ").strip().lower().split())


def find_column(columns: list[str], aliases: list[str]) -> str | None:
    normalized: dict[str, str] = {}
    for column in columns:
        normalized.setdefault(normalize_header(column), column)
    for alias in aliases:
        match = normalized.get(normalize_header(alias))
        if match is not None:
            return match
    return None


def validate_master_data_columns(read_result: ReadResult, fields_to_audit: list[str]) -> None:
    columns = read_result.columns
    missing_required: list[str] = []
    for label, aliases in (
        ("Brand", CANONICAL_ALIASES["brand"]),
        ("STATUS", CANONICAL_ALIASES["status"]),
        ("SKU / item code", CANONICAL_ALIASES["sku"]),
        ("Product name", CANONICAL_ALIASES["product_name"]),
    ):
        if find_column(columns, aliases) is None:
            missing_required.append(label)

    available_audit_fields = [
        field
        for field in fields_to_audit
        if find_column(columns, FIELD_ALIASES.get(field, [field])) is not None
    ]
    if not missing_required and available_audit_fields:
        return

    parts: list[str] = []
    if missing_required:
        parts.append(f"Missing required column(s): {', '.join(missing_required)}")
    if not available_audit_fields:
        examples = ", ".join(fields_to_audit[:8])
        parts.append(f"No DQC audit fields were found. Expected fields include: {examples}")
    parts.append("Please upload the original master data file for Data Quality Control.")
    raise ValueError(" | ".join(parts))


def read_master_data(path: Path, fields_to_audit: list[str]) -> ReadResult:
    if not path.exists():
        raise FileNotFoundError(f"Input file not found: {path}")
    if path.suffix.lower() not in {".xlsx", ".xlsm", ".csv"}:
        raise ValueError("Input file must be .xlsx, .xlsm, or .csv")

    if path.suffix.lower() == ".csv":
        df = pd.read_csv(path, dtype=str, keep_default_na=False)
        df = _clean_dataframe(df)
        result = ReadResult(df, "csv", 1, path, list(df.columns), [])
        validate_master_data_columns(result, fields_to_audit)
        return result

    sheet_name, header_row, warnings = detect_excel_layout(path, fields_to_audit)
    df = pd.read_excel(path, sheet_name=sheet_name, header=header_row - 1, dtype=str, keep_default_na=False)
    df = _clean_dataframe(df)
    if df.empty:
        raise ValueError("No data rows were found after the detected header row")
    result = ReadResult(df, sheet_name, header_row, path, list(df.columns), warnings)
    validate_master_data_columns(result, fields_to_audit)
    return result


def detect_excel_layout(path: Path, fields_to_audit: list[str]) -> tuple[str, int, list[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    warnings: list[str] = []
    generated_report_sheets = {"Missing Data Overview", "Summary Tracker", "Action Tracker", "Run Summary"}
    if generated_report_sheets.intersection(set(workbook.sheetnames)):
        warnings.append("Input appears to be a generated audit report. Use the original master data file for accurate results.")
    known_headers = {
        normalize_header(alias)
        for aliases in CANONICAL_ALIASES.values()
        for alias in aliases
    }
    known_headers.update(normalize_header(field) for field in fields_to_audit)

    best: tuple[int, str, int] | None = None
    for worksheet in workbook.worksheets:
        max_row = min(worksheet.max_row or 1, 40)
        max_col = min(worksheet.max_column or 1, 120)
        for row_idx in range(1, max_row + 1):
            values = [normalize_header(worksheet.cell(row_idx, col_idx).value) for col_idx in range(1, max_col + 1)]
            non_empty = [value for value in values if value]
            if not non_empty:
                continue
            score = sum(1 for value in non_empty if value in known_headers)
            has_brand = any(value in {normalize_header(alias) for alias in CANONICAL_ALIASES["brand"]} for value in non_empty)
            if has_brand:
                score += 5
            if score > 0 and (best is None or score > best[0]):
                best = (score, worksheet.title, row_idx)

    if best is None:
        examples = ", ".join(["Brand", "STATUS", "SKU", *fields_to_audit[:5]])
        raise ValueError(
            "Could not detect a valid master data header row. "
            f"Expected columns include: {examples}. "
            "Please upload the original master data file for Data Quality Control."
        )
    return best[1], best[2], warnings


def _clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(column).strip() if str(column).strip() else f"Unnamed {idx}" for idx, column in enumerate(df.columns, 1)]
    df = df.dropna(how="all")
    df = df.loc[:, ~pd.Index(df.columns).duplicated()]
    for column in df.columns:
        df[column] = df[column].map(lambda value: "" if value is None else str(value).strip())
    return df.reset_index(drop=True)
