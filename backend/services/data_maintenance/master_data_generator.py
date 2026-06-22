"""Master Data Generator — produces [Brand]_Missing_Data and [Brand]_Missing_Data_Status Excel files."""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

STATE_FILE_NAME = "master_data_state.json"
VALID_STATUSES = ["Active", "Upcoming", "Limited", "Non-ACR", "Blanks", "N/A", "Unknown", "Others"]
_MASTER_DATA_CACHE: dict[tuple[str, int, int], pd.DataFrame] = {}

TEMPLATE_HEADERS_ROW = 12
TEMPLATE_DATA_START = 13

SECTION_GROUPS = [
    ("B", "Basic product information"),
    ("D", "Consumer pack (Size in mm)"),
    ("J", "Trade pack/Inner box (Size in mm)"),
    ("O", "Outer Box (Size in mm)"),
    ("T", "Registration"),
    ("AA", "Product compliance and labeing"),
    ("AM", "Product price"),
]

COLUMN_HEADERS = [
    "OT\nArticle\nNumber",
    "BarCode",
    "Product name \n(If applicable, include sub-brand)",
    "Volume/Unit",
    "Net Weight\n (g)",
    "Gross weight\n (g)",
    "DEPTH /\nLENGTH",
    "HEIGHT",
    "WIDTH",
    "Inner box QTY\n[Mark \"X\" if there is no inner box for the item]",
    "Gross weight \n(g)",
    "DEPTH /\nLENGTH",
    "HEIGHT",
    "WIDTH",
    "Outer box QTY",
    "Gross weight\n(g)",
    "DEPTH /\nLENGTH",
    "HEIGHT",
    "WIDTH",
    "CPNP Number",
    "UK \nSCPN NUMBER",
    "Manufacturer name",
    "Manufacturer address",
    "Manufacturer URL",
    "EU Responsible person\n(name, address, email address)",
    "UK Responsible person\n(name, address, email address)",
    "Ingredient list",
    "PAO\n(Months)",
    "Shelf Life\n(Months)",
    "Description\n (250+ words)",
    "Instructions/How to use",
    "Warninngs",
    "Vegan\n(Y/N)",
    "Is EXP date on pack mentioned?\n(Y/N)",
    "Is PAO mentioned on Pack?\n(Y/N)",
    "Is it compliant with Directive 2024/825 (EU) on green claims? \n(Y/N/NA)",
    "Cruelty\nFree?\n(Y/N)",
    "#HEX CODE",
    "SUPPLY PRICE",
    "SUPPLY CURRENCY",
    "RECOMMENDED RRP",
    "Comment if Renewed: What Has Changed",
]

MASTER_TO_TEMPLATE = [
    ("CODE", 1),
    ("BAR CODE", 2),
    ("NAME", 3),
    ("UNIT/VOLUME", 4),
    ("WEIGHT", 5),
    ("GROSS WEIGHT", 6),
    ("DEPTH", 7),
    ("HEIGHT", 8),
    ("WIDTH", 9),
    ("INNER BOX QTY", 10),
    ("INNER BOX G.WEIGHT", 11),
    ("INNER BOX LENGTH", 12),
    ("INNER BOX HEIGHT", 13),
    ("INNER BOX WIDTH", 14),
    ("CARTON QTY", 15),
    ("OUTER BOX G.WEIGHT", 16),
    ("OUTER BOX LENGTH", 17),
    ("OUTER BOX HEIGHT", 18),
    ("OUTER BOX WIDTH", 19),
    ("CPNP NUMBER", 20),
    ("SCPN NUMBER", 21),
    ("MANUFACTURER INFORMATION", 22),
    (None, 23),
    (None, 24),
    ("EU RESPONSIBLE PERSON", 25),
    ("UK RESPONSIBLE PERSON", 26),
    ("INGREDIENTS", 27),
    ("PAO", 28),
    ("SHELF LIFE", 29),
    ("DESCRIPTION", 30),
    (None, 31),
    ("WARNING", 32),
    ("VEGAN CERTIFIED?", 33),
    (None, 34),
    (None, 35),
    (None, 36),
    (None, 37),
    (None, 38),
    ("SUPPLY PRICE", 39),
    ("SUPPLY PRICE CURRENCY", 40),
    (None, 41),
    (None, 42),
]


def _safe_val(row: pd.Series, col: str | None) -> Any:
    if col is None:
        return None
    val = row.get(col)
    if pd.isna(val):
        return None
    if isinstance(val, float) and val == 0.0:
        return 0.0
    return val


def read_dqc_brands(dqc_path: Path) -> list[str]:
    try:
        xls = pd.ExcelFile(dqc_path)
        if "SKU Missing Detail" in xls.sheet_names:
            df = pd.read_excel(xls, "SKU Missing Detail")
            if "Brand" in df.columns:
                return sorted(df["Brand"].dropna().unique().tolist())
        if "Missing Data Overview" in xls.sheet_names:
            df = pd.read_excel(xls, "Missing Data Overview", header=4)
            if "Brand" in df.columns:
                return sorted(df["Brand"].dropna().unique().tolist())
    except Exception:
        pass
    return []


def read_master_data(master_path: Path) -> pd.DataFrame:
    path = master_path.resolve()
    stat = path.stat()
    key = (str(path), int(stat.st_mtime_ns), int(stat.st_size))
    cached = _MASTER_DATA_CACHE.get(key)
    if cached is not None:
        return cached
    xls = pd.ExcelFile(path)
    try:
        df = pd.read_excel(xls, xls.sheet_names[0])
    finally:
        xls.close()
    _MASTER_DATA_CACHE.clear()
    _MASTER_DATA_CACHE[key] = df
    return df


def get_brand_products(master_df: pd.DataFrame, brand: str) -> pd.DataFrame:
    if "OBJECT" not in master_df.columns:
        return pd.DataFrame()
    return master_df[master_df["OBJECT"].str.upper() == brand.upper()].copy()


def generate_missing_data(products: pd.DataFrame, brand: str, output_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Product info template V03"

    header_font = Font(name="Aptos Narrow", size=11, bold=True)
    data_font = Font(name="Aptos Narrow", size=11)
    red_fill = PatternFill("solid", fgColor="FFC00000")
    blue_fill = PatternFill("solid", fgColor="FF2E75B6")
    white_font = Font(name="Aptos Narrow", size=11, bold=True, color="FFFFFF")
    section_font = Font(name="Aptos Narrow", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    ws.cell(1, 1, "Production Information Template").font = Font(name="Aptos Narrow", size=14, bold=True)
    ws.cell(3, 1, "1. Please fill in the minimum required data (coloured red) for a valid first price offer to retailers.").font = data_font
    ws.cell(4, 1, "2. Please provide the packshot and artwork image files separately (email or drive link), with barcode in the image file name.").font = data_font
    ws.cell(5, 1, "3. All additional information (coloured blue) is required for order and activation-ready status of the product.").font = data_font

    ws.cell(8, 1, "Minimum required for first price offer").font = Font(name="Aptos Narrow", size=11, bold=True, color="C00000")
    ws.cell(9, 1, "Required for ordering").font = Font(name="Aptos Narrow", size=11, bold=True, color="2E75B6")
    ws.cell(9, 39, "Kindly provide an update if requested.").font = data_font

    from openpyxl.utils import get_column_letter
    section_cols = {"B": "Basic product information", "D": "Consumer pack (Size in mm)",
                    "J": "Trade pack/Inner box (Size in mm)", "O": "Outer Box (Size in mm)",
                    "T": "Registration", "AA": "Product compliance and labeing", "AM": "Product price"}
    for col_letter, label in section_cols.items():
        from openpyxl.utils import column_index_from_string
        col_idx = column_index_from_string(col_letter)
        cell = ws.cell(11, col_idx, label)
        cell.font = section_font
        cell.fill = PatternFill("solid", fgColor="FF4472C4")

    red_cols = {1, 2, 3, 4, 5, 6, 7, 8, 9, 15, 39, 40, 41}
    blue_cols = set(range(10, 15)) | set(range(16, 20)) | set(range(20, 39)) | {42}

    for col_idx, header_text in enumerate(COLUMN_HEADERS, 1):
        cell = ws.cell(TEMPLATE_HEADERS_ROW, col_idx, header_text)
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        if col_idx in red_cols:
            cell.fill = red_fill
        elif col_idx in blue_cols:
            cell.fill = blue_fill
        else:
            cell.fill = PatternFill("solid", fgColor="FF4472C4")

    for col_idx in range(1, 43):
        col_letter = get_column_letter(col_idx)
        if col_idx in {3, 25, 26, 27, 30, 31}:
            ws.column_dimensions[col_letter].width = 30
        elif col_idx in {1, 2, 22, 23, 24}:
            ws.column_dimensions[col_letter].width = 18
        elif col_idx in {39, 40, 41, 42}:
            ws.column_dimensions[col_letter].width = 16
        else:
            ws.column_dimensions[col_letter].width = 13

    for row_offset, (_, product_row) in enumerate(products.iterrows()):
        excel_row = TEMPLATE_DATA_START + row_offset
        for master_col, template_col in MASTER_TO_TEMPLATE:
            val = _safe_val(product_row, master_col)
            if val is not None:
                cell = ws.cell(excel_row, template_col, val)
                cell.font = data_font
                cell.border = thin_border

    filename = f"{brand}_Missing_Data.xlsx"
    file_path = output_path / filename
    wb.save(file_path)
    return file_path


def generate_missing_data_status(products: pd.DataFrame, brand: str, output_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    header_font = Font(name="Aptos Narrow", size=11, bold=True)
    data_font = Font(name="Aptos Narrow", size=11)
    header_fill = PatternFill("solid", fgColor="FF4472C4")
    white_font = Font(name="Aptos Narrow", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    headers = ["OT ARTICLE NUMBER", "BARCODE", "STATUS", "PRODUCT NAME"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(1, col_idx, header)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 55

    for row_offset, (_, product_row) in enumerate(products.iterrows()):
        excel_row = 2 + row_offset
        code = _safe_val(product_row, "CODE")
        barcode = _safe_val(product_row, "BAR CODE")
        status = _safe_val(product_row, "STATUS")
        name = _safe_val(product_row, "NAME")

        for col_idx, val in enumerate([code, barcode, status, name], 1):
            cell = ws.cell(excel_row, col_idx, val)
            cell.font = data_font
            cell.border = thin_border

    ws2 = wb.create_sheet("Sheet2")
    ws2.cell(1, 1, "STATUS").font = header_font
    ws2.column_dimensions["A"].width = 14
    for i, status in enumerate(VALID_STATUSES, 2):
        ws2.cell(i, 1, status).font = data_font

    filename = f"{brand}_Missing_Data_Status.xlsx"
    file_path = output_path / filename
    wb.save(file_path)
    return file_path


def save_state(state_dir: Path, state: dict) -> None:
    state_dir.mkdir(parents=True, exist_ok=True)
    state_file = state_dir / STATE_FILE_NAME
    state_file.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")


def load_state(state_dir: Path) -> dict | None:
    state_file = state_dir / STATE_FILE_NAME
    if not state_file.exists():
        return None
    try:
        return json.loads(state_file.read_text(encoding="utf-8"))
    except Exception:
        return None
