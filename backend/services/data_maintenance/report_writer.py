from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .auditor import AuditResult


SHEETS = [
    ("Executive Summary", "executive_summary"),
    ("Brand Scorecard", "brand_scorecard"),
    ("Missing Data Overview", "missing_overview"),
    ("Summary Tracker", "summary_tracker"),
    ("Action Tracker", "action_tracker"),
    ("SKU Missing Detail", "sku_missing_detail"),
    ("Validation Errors", "validation_errors"),
    ("Run Summary", "run_summary"),
]


def write_audit_report(result: AuditResult, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, attr_name in SHEETS:
            dataframe: pd.DataFrame = getattr(result, attr_name)
            if dataframe.empty:
                dataframe = pd.DataFrame({"Message": ["No rows"]})
            dataframe.to_excel(writer, index=False, sheet_name=sheet_name)

    format_workbook(output_path)
    return output_path


def format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for column_cells in worksheet.columns:
            max_length = 0
            column_index = column_cells[0].column
            for cell in column_cells[:200]:
                max_length = max(max_length, len(str(cell.value or "")))
            worksheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 12), 48)
    workbook.save(path)
